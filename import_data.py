#!/usr/bin/env python
"""批量导入：取得发票 + 进项抵扣 数据到数据库"""
import os, sys, hashlib, json, glob, re
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from database import SessionLocal, PurchaseInvoice, InputVATDeduction

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "tax-risk-docs")
COMPANY_ID = 2


def _safe_float(v, default=0.0):
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(",", "").replace("￥", "").replace("¥", "")
    try: return float(s)
    except: return default


def _parse_date(s, fmts=None):
    if not s: return None
    s = str(s).strip()
    if fmts is None:
        fmts = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d",
                "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"]
    for f in fmts:
        try: return datetime.strptime(s, f).date()
        except: pass
    return None


def import_purchase_invoices():
    """导入所有取票文件"""
    db = SessionLocal()
    imported = 0
    skipped = 0
    errors = 0

    files = sorted(glob.glob(os.path.join(UPLOAD_DIR, f"{COMPANY_ID}_*取票*.xlsx")))
    print(f"找到 {len(files)} 个取票文件")

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb.active
            headers = []
            for c in range(1, ws.max_column + 1):
                headers.append(str(ws.cell(row=1, column=c).value or "").strip())

            # 列映射：表头 → 字段
            col_map = {}
            hdr_map = {
                "发票代码": "invoice_code", "发票号码": "invoice_no", "数电发票号码": "digital_invoice_no",
                "销方识别号": "seller_tax_no", "销方名称": "seller_name",
                "购方识别号": "buyer_tax_no", "购买方名称": "buyer_name",
                "开票日期": "invoice_date",
                "税收分类编码": "tax_category_code", "特定业务类型": "specific_business_type",
                "货物或应税劳务名称": "goods_name", "规格型号": "spec", "单位": "unit",
                "数量": "quantity", "单价": "unit_price",
                "金额": "amount", "税率": "tax_rate", "税额": "tax_amount",
                "价税合计": "total_amount",
                "发票来源": "invoice_source", "发票票种": "invoice_category",
                "发票状态": "status", "是否正数发票": "is_positive",
                "发票风险等级": "invoice_risk_level", "开票人": "issuer", "备注": "remark",
            }
            for i, h in enumerate(headers):
                if h in hdr_map:
                    col_map[hdr_map[h]] = i

            for r in range(2, ws.max_row + 1):
                vals = {}
                for field, ci in col_map.items():
                    v = ws.cell(row=r, column=ci + 1).value
                    vals[field] = str(v).strip() if v is not None else ""

                # 跳过完全空行
                if not any(v for v in vals.values()):
                    continue

                inv_no = vals.get("invoice_no", "").strip()
                inv_date = _parse_date(vals.get("invoice_date", ""))
                qty = _safe_float(vals.get("quantity", 0))
                uprice = _safe_float(vals.get("unit_price", 0))
                amt = _safe_float(vals.get("amount", 0))
                tr = _safe_float(vals.get("tax_rate", 0))
                tax_amt = _safe_float(vals.get("tax_amount", 0))
                total = _safe_float(vals.get("total_amount", 0))

                # 全指纹去重
                fp_values = (
                    str(COMPANY_ID), str(inv_no or ""), str(vals.get("invoice_code", "")),
                    str(vals.get("digital_invoice_no", "")),
                    str(vals.get("seller_tax_no", "")), str(vals.get("seller_name", "")),
                    str(vals.get("buyer_tax_no", "")), str(vals.get("buyer_name", "")),
                    str(inv_date) if inv_date else "",
                    str(vals.get("tax_category_code", "")), str(vals.get("specific_business_type", "")),
                    str(vals.get("goods_name", "")), str(vals.get("spec", "")),
                    str(vals.get("unit", "")), str(qty), str(uprice),
                    str(amt), str(tr), str(tax_amt), str(total),
                    str(vals.get("invoice_source", "")),
                    str(vals.get("invoice_category", "增值税专用发票")),
                    str(vals.get("status", "正常")),
                    str(vals.get("is_positive", "是")),
                    str(vals.get("invoice_risk_level", "")),
                    str(vals.get("issuer", "")),
                    str(vals.get("remark", "")),
                )
                fp_raw = "|".join(fp_values)
                fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()

                existing = db.query(PurchaseInvoice).filter(
                    PurchaseInvoice.company_id == COMPANY_ID,
                    PurchaseInvoice._fingerprint == fp
                ).first()
                if existing:
                    skipped += 1
                    continue

                inv = PurchaseInvoice(
                    company_id=COMPANY_ID, invoice_no=inv_no,
                    invoice_code=vals.get("invoice_code", ""),
                    digital_invoice_no=vals.get("digital_invoice_no", ""),
                    seller_tax_no=vals.get("seller_tax_no", ""),
                    seller_name=vals.get("seller_name", ""),
                    buyer_tax_no=vals.get("buyer_tax_no", ""),
                    buyer_name=vals.get("buyer_name", ""),
                    invoice_date=inv_date,
                    tax_category_code=vals.get("tax_category_code", ""),
                    specific_business_type=vals.get("specific_business_type", ""),
                    goods_name=vals.get("goods_name", ""),
                    spec=vals.get("spec", ""),
                    unit=vals.get("unit", ""),
                    quantity=qty, unit_price=uprice,
                    amount=amt, tax_rate=tr, tax_amount=tax_amt,
                    total_amount=total,
                    invoice_source=vals.get("invoice_source", ""),
                    invoice_category=vals.get("invoice_category", "增值税专用发票"),
                    status=vals.get("status", "正常"),
                    is_positive=vals.get("is_positive", "是") in ("是", "true", "True", "1", True),
                    invoice_risk_level=vals.get("invoice_risk_level", ""),
                    issuer=vals.get("issuer", ""),
                    remark=vals.get("remark", ""),
                    _fingerprint=fp,
                )
                db.add(inv)
                imported += 1

                if imported % 500 == 0:
                    db.commit()
                    print(f"  已导入 {imported}...")

            db.commit()
            print(f"  {fname}: 导入 {imported} 条 (累计), 跳过 {skipped} 条重复")

        except Exception as e:
            errors += 1
            print(f"  {fname}: 错误 - {e}")
            db.rollback()

    db.commit()
    print(f"\n取得发票导入完成: {imported} 条新记录, {skipped} 条跳过, {errors} 个错误")
    db.close()
    return imported


def import_input_vat_deductions():
    """导入所有抵扣文件"""
    db = SessionLocal()
    imported = 0
    skipped = 0
    errors = 0

    files = sorted(glob.glob(os.path.join(UPLOAD_DIR, f"{COMPANY_ID}_*抵扣*.xlsx")))
    print(f"找到 {len(files)} 个抵扣文件")

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb.active
            headers = []
            for c in range(1, ws.max_column + 1):
                headers.append(str(ws.cell(row=1, column=c).value or "").strip())

            hdr_map = {
                "勾选状态": "check_status", "发票来源": "invoice_source",
                "转内销证明编号": "domestic_sale_cert_no", "数电发票号码": "digital_invoice_no",
                "发票代码": "invoice_code", "发票号码": "invoice_no",
                "开票日期": "invoice_date",
                "销售方纳税人识别号": "seller_tax_id", "销售方纳税人名称": "seller_name",
                "金额": "amount", "税额": "tax_amount", "有效抵扣税额": "deductible_tax_amount",
                "票种": "invoice_category", "票种标签": "invoice_category_label",
                "发票状态": "invoice_status", "勾选时间": "check_time",
                "发票风险等级": "risk_level",
            }
            col_map = {}
            for i, h in enumerate(headers):
                if h in hdr_map:
                    col_map[hdr_map[h]] = i

            import_batch = f"{fname}_{datetime.now().timestamp()}"

            for r in range(2, ws.max_row + 1):
                vals = {}
                for field, ci in col_map.items():
                    v = ws.cell(row=r, column=ci + 1).value
                    vals[field] = str(v).strip() if v is not None else ""

                if not any(v for v in vals.values()):
                    continue

                inv_date = _parse_date(vals.get("invoice_date", ""))
                check_time_str = vals.get("check_time", "")
                check_time = None
                if check_time_str:
                    for f in ["%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                              "%Y/%m/%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]:
                        try:
                            check_time = datetime.strptime(check_time_str, f)
                            break
                        except: pass

                amt = _safe_float(vals.get("amount", 0))
                tax_amt = _safe_float(vals.get("tax_amount", 0))
                deductible = _safe_float(vals.get("deductible_tax_amount", 0))

                # 全行指纹
                fp_values = (
                    str(COMPANY_ID),
                    str(vals.get("check_status", "")),
                    str(vals.get("invoice_source", "")),
                    str(vals.get("domestic_sale_cert_no", "")),
                    str(vals.get("digital_invoice_no", "")),
                    str(vals.get("invoice_code", "")),
                    str(vals.get("invoice_no", "")),
                    str(inv_date) if inv_date else "",
                    str(vals.get("seller_tax_id", "")),
                    str(vals.get("seller_name", "")),
                    str(amt), str(tax_amt), str(deductible),
                    str(vals.get("invoice_category", "")),
                    str(vals.get("invoice_category_label", "")),
                    str(vals.get("invoice_status", "正常")),
                    str(check_time) if check_time else "",
                    str(vals.get("risk_level", "正常")),
                )
                fp_raw = "|".join(fp_values)
                fp = hashlib.sha256(fp_raw.encode("utf-8")).hexdigest()

                existing = db.query(InputVATDeduction).filter(
                    InputVATDeduction.company_id == COMPANY_ID,
                    InputVATDeduction._fingerprint == fp
                ).first()
                if existing:
                    skipped += 1
                    continue

                ded = InputVATDeduction(
                    company_id=COMPANY_ID,
                    check_status=vals.get("check_status", ""),
                    invoice_source=vals.get("invoice_source", ""),
                    domestic_sale_cert_no=vals.get("domestic_sale_cert_no", ""),
                    digital_invoice_no=vals.get("digital_invoice_no", ""),
                    invoice_code=vals.get("invoice_code", ""),
                    invoice_no=vals.get("invoice_no", ""),
                    invoice_date=inv_date,
                    seller_tax_id=vals.get("seller_tax_id", ""),
                    seller_name=vals.get("seller_name", ""),
                    amount=amt, tax_amount=tax_amt,
                    deductible_tax_amount=deductible,
                    invoice_category=vals.get("invoice_category", ""),
                    invoice_category_label=vals.get("invoice_category_label", ""),
                    invoice_status=vals.get("invoice_status", "正常"),
                    check_time=check_time,
                    risk_level=vals.get("risk_level", "正常"),
                    import_batch_id=import_batch,
                    _fingerprint=fp,
                )
                db.add(ded)
                imported += 1

                if imported % 500 == 0:
                    db.commit()
                    print(f"  已导入 {imported}...")

            db.commit()
            print(f"  {fname}: 导入 {imported} 条 (累计)")

        except Exception as e:
            errors += 1
            print(f"  {fname}: 错误 - {e}")
            import traceback
            traceback.print_exc()
            db.rollback()

    db.commit()
    print(f"\n进项抵扣导入完成: {imported} 条新记录, {skipped} 条跳过, {errors} 个错误")
    db.close()
    return imported


if __name__ == "__main__":
    print("=" * 60)
    print("开始导入数据...")
    print("=" * 60)
    n1 = import_purchase_invoices()
    n2 = import_input_vat_deductions()
    print(f"\n总计: 取得发票 {n1} 条 + 进项抵扣 {n2} 条")
