"""
税务稽查分析脚本 — 以稽查员身份全面审计23个文件
支持多格式银行流水（含多行头表）和发票
"""
import os, sys, json, re
from collections import defaultdict, Counter
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "tax-risk-docs")

def read_excel_rows(fpath):
    """读取 Excel 返回 (header_cols, data_rows)"""
    header_row = []
    data_rows = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_list = []
            for row in ws.iter_rows():
                vals = [str(c.value).strip() if c.value is not None else '' for c in row]
                rows_list.append(vals)
            # 找表头行(含"交易时间"或"收入金额"或"对方"的行)
            header_idx = 0
            for i, r in enumerate(rows_list[:10]):
                row_text = ' '.join(r)
                if any(kw in row_text for kw in ['交易时间', '对方户名', '收入金额', '交易日期', '对方名称', '发票号码', '货物或应税劳务名称', '销方名称', '购方名称', '金额', '价税合计']):
                    header_idx = i
                    break
            header_row = rows_list[header_idx] if header_idx < len(rows_list) else []
            data_rows.extend(rows_list[header_idx+1:])
    except:
        try:
            import xlrd
            wb = xlrd.open_workbook(fpath)
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                rows_list = []
                for r in range(ws.nrows):
                    vals = [str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) != '' else '' for c in range(ws.ncols)]
                    rows_list.append(vals)
                # 找表头行
                header_idx = 0
                for i, r in enumerate(rows_list[:10]):
                    row_text = ' '.join(r)
                    if any(kw in row_text for kw in ['交易时间', '对方户名', '收入金额', '交易日期', '对方名称', '发票号码', '货物或应税劳务名称', '销方名称', '购方名称', '金额', '价税合计']):
                        header_idx = i
                        break
                header_row = rows_list[header_idx] if header_idx < len(rows_list) else []
                data_rows.extend(rows_list[header_idx+1:])
        except Exception as e2:
            print(f"  ERROR: {os.path.basename(fpath)}: {e2}")
    return header_row, data_rows


def detect_type(header_row):
    """根据表头检测文件类型"""
    row_text = ' '.join(h.lower() for h in header_row)
    inv_kw = ['货物或应税劳务', 'goods', '销方', '购方', '发票代码', '发票号码', '金额', '税额', '价税合计', 'seller', 'buyer', 'inv_no']
    bank_kw = ['交易时间', '交易日期', '对方户名', '交易对手', 'counterparty', '收入金额', '支出金额', '本次余额', '借贷', '借方', '贷方', '余额', '对方账号', '交易行名', '交易类型', '对方名称', '对方省市']
    
    inv_score = sum(1 for kw in inv_kw if kw in row_text)
    bank_score = sum(1 for kw in bank_kw if kw in row_text)
    
    if bank_score >= 3:
        return "银行流水"
    elif inv_score >= 3:
        return "发票"
    else:
        return "其他"


def parse_invoices(header_row, data_rows, company_name=''):
    """解析发票数据，返回 [{...}, ...]
    company_name: 被查单位名称，用于判定发票方向（销项=本公司是销售方，进项=本公司是购买方）
    """
    # 建立列索引
    col_map = {}
    for i, h in enumerate(header_row):
        h = h.lower()
        if '销方' in h or 'seller' in h or '销售方' in h: col_map['seller'] = i
        if '购方' in h or 'buyer' in h or '购买方' in h: col_map['buyer'] = i
        if '货物' in h or 'goods' in h or '品名' in h: col_map['goods'] = i
        if '发票号码' in h or 'inv_no' in h or '发票号' in h: col_map['inv_no'] = i
        if '价税合计' in h or ('合计' in h and '金额' in h): col_map['total'] = i
        elif '金' in h and '额' in h and '税' not in h and '合' not in h and '价' not in h: col_map['amount'] = i
        if '税' in h and '额' in h: col_map['tax'] = i
        if '数量' in h: col_map['qty'] = i
        if '单价' in h: col_map['price'] = i
        if '金额' in h and '税' not in h and '合' not in h and '价' not in h and '数量' not in h:
            if 'amount' not in col_map or '不含' in h: col_map['amount'] = i
    
    invoices = []
    for row in data_rows:
        if not row or all(not c for c in row): continue
        
        inv = {}
        for key, idx in col_map.items():
            if idx < len(row) and row[idx]:
                inv[key] = row[idx]
        
        if not inv: continue
        
        # 数值转换
        for k in ['total', 'amount', 'tax', 'qty', 'price']:
            if k in inv:
                v = inv[k]
                if isinstance(v, str):
                    try: inv[k] = float(v.replace(',', '').replace('¥', '').replace(' ', ''))
                    except: del inv[k]
        
        # 方向判定 — 用公司名称判断，不硬编码
        buyer = str(inv.get('buyer', ''))
        seller = str(inv.get('seller', ''))
        
        if company_name and company_name in seller:
            inv['direction'] = '销项'
        elif company_name and company_name in buyer:
            inv['direction'] = '进项'
        else:
            # 公司名未提供或找不到 → 按发票字段模糊推断
            if seller and not buyer:
                inv['direction'] = '进项'
            elif buyer and not seller:
                inv['direction'] = '销项'
            else:
                inv['direction'] = '未知'
        
        invoices.append(inv)
    return invoices


def parse_bank_txs(header_row, data_rows):
    """解析银行流水，返回 [{...}, ...]"""
    col_map = {}
    for i, h in enumerate(header_row):
        h = h.lower().replace(' ', '')
        if '交易时间' in h or '交易日期' in h or 'date' in h or 'tx_time' in h:
            col_map['date'] = i
        if '对方户名' in h or '交易对手' in h or 'counterparty' in h or '对方名称' in h:
            col_map['counterparty'] = i
        if '收入金额' in h or ('收入' in h and '金' in h) or 'credit' in h or '贷方' in h:
            col_map['credit'] = i
        if '支出金额' in h or ('支出' in h and '金' in h) or 'debit' in h or '借方' in h:
            col_map['debit'] = i
        if '余额' in h or 'balance' in h: col_map['balance'] = i
        if '交易用途' in h or '摘要' in h or 'summary' in h or '用途' in h: col_map['summary'] = i
        if '对方账号' in h: col_map['counterparty_acct'] = i
        if '对方省市' in h: col_map['counterparty_city'] = i
        if '交易行名' in h: col_map['bank_branch'] = i
    
    txs = []
    for row in data_rows:
        if not row or all(not c for c in row): continue
        if len(row[0].strip()) < 5: continue  # skip too short date
        
        tx = {}
        for key, idx in col_map.items():
            if idx < len(row) and row[idx]:
                tx[key] = row[idx]
        
        if not tx.get('date') and not tx.get('counterparty'):
            continue
        
        # 数值转换
        for k in ['credit', 'debit', 'balance']:
            if k in tx:
                v = tx[k]
                if isinstance(v, str):
                    try: tx[k] = float(v.replace(',', '').replace('¥', '').replace(' ', ''))
                    except: tx[k] = 0
                elif not isinstance(v, (int, float)):
                    tx[k] = 0
        
        # 过滤汇总行：大额整数+无对手+摘要含"汇总"或"摘要"
        summary = str(tx.get('summary', ''))
        debit = safe_float(tx.get('debit'))
        credit = safe_float(tx.get('credit'))
        cp = str(tx.get('counterparty', '')).strip()
        
        if (debit > 100000 or credit > 100000) and not cp and ('汇总' in summary or '小计' in summary):
            continue
        
        txs.append(tx)
    return txs


def safe_float(v):
    """安全转换为float"""
    if v is None: return 0
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace(',', '').replace('¥', '').replace(' ', ''))
        except: return 0
    return 0

# ════════════════ MAIN ════════════════

print("=" * 70)
print("税务稽查分析报告")
print(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
print("=" * 70)

files = sorted(os.listdir(UPLOAD_DIR))
all_invoices = []
all_bank_txs = []

for fname in files:
    fpath = os.path.join(UPLOAD_DIR, fname)
    header_row, data_rows = read_excel_rows(fpath)
    ftype = detect_type(header_row)
    
    if ftype == "发票":
        invs = parse_invoices(header_row, data_rows)
        for i in invs: i['_source'] = fname
        all_invoices.extend(invs)
        print(f"[发票] {fname}: {len(invs)} 条")
    elif ftype == "银行流水":
        txs = parse_bank_txs(header_row, data_rows)
        for t in txs: t['_source'] = fname
        all_bank_txs.extend(txs)
        print(f"[银行] {fname}: {len(txs)} 条")
    else:
        # 尝试强制解析：看表头是否含"交易"或"金额"
        htext = ' '.join(str(h) for h in header_row)
        if '交易时间' in htext or '对方户名' in htext:
            txs = parse_bank_txs(header_row, data_rows)
            if txs:
                for t in txs: t['_source'] = fname
                all_bank_txs.extend(txs)
                print(f"[银行*] {fname}: {len(txs)} 条")
            else:
                print(f"[跳过] {fname}: 未识别")
        else:
            print(f"[跳过] {fname}: {htext[:60]}...")

# 分离进销项
pur_invs = [i for i in all_invoices if i.get('direction') == '进项']
sal_invs = [i for i in all_invoices if i.get('direction') == '销项']

print(f"\n总计: 进项{len(pur_invs)}张, 销项{len(sal_invs)}张, 银行流水{len(all_bank_txs)}条")

# ═══════ 银行业务分析 ═══════
total_in = sum(safe_float(tx.get('credit')) for tx in all_bank_txs)
total_out = sum(safe_float(tx.get('debit')) for tx in all_bank_txs)

print(f"\n银行流水: 收款{total_in:,.0f}元, 付款{total_out:,.0f}元, 净现金流{total_in-total_out:,.0f}元")

# 收款构成
credit_by_cp = defaultdict(float)
for tx in all_bank_txs:
    c = safe_float(tx.get('credit'))
    if c > 0:
        cp = str(tx.get('counterparty', '')).strip()
        if cp: credit_by_cp[cp] += c

# 付款构成
debit_by_cp = defaultdict(float)
for tx in all_bank_txs:
    d = safe_float(tx.get('debit'))
    if d > 0:
        cp = str(tx.get('counterparty', '')).strip()
        if cp: debit_by_cp[cp] += d

# ═══════ 进项分析 ═══════
total_pur = sum(safe_float(i.get('total', i.get('amount'))) for i in pur_invs)
print(f"\n进项: {len(pur_invs)}张, 采购{total_pur:,.0f}元")

# 供应商分析
supplier_stats = defaultdict(lambda: {"amt": 0, "cnt": 0, "goods": set(), "city": ""})
for i in pur_invs:
    s = str(i.get('seller', ''))[:25]
    a = safe_float(i.get('total', i.get('amount')))
    g = str(i.get('goods', ''))[:25]
    supplier_stats[s]['amt'] += a
    supplier_stats[s]['cnt'] += 1
    supplier_stats[s]['goods'].add(g)

# 品类分析
pur_goods_amt = defaultdict(float)
for i in pur_invs:
    g = str(i.get('goods', ''))[:20]
    a = safe_float(i.get('total', i.get('amount')))
    pur_goods_amt[g] += a

# ═══════ 销项分析 ═══════
total_sal = sum(safe_float(i.get('total', i.get('amount'))) for i in sal_invs)
print(f"销项: {len(sal_invs)}张, 销售{total_sal:,.0f}元")

# 客户分析
buyer_stats = defaultdict(lambda: {"amt": 0, "cnt": 0, "goods": set()})
for i in sal_invs:
    b = str(i.get('buyer', ''))[:25]
    a = safe_float(i.get('total', i.get('amount')))
    g = str(i.get('goods', ''))[:25]
    buyer_stats[b]['amt'] += a
    buyer_stats[b]['cnt'] += 1
    buyer_stats[b]['goods'].add(g)

# 销项品类
sal_goods_amt = defaultdict(float)
for i in sal_invs:
    g = str(i.get('goods', ''))[:20]
    a = safe_float(i.get('total', i.get('amount')))
    sal_goods_amt[g] += a

# ═══════ 交叉比对 ═══════
# 进项发票供应商 vs 银行付款记录
bank_payees = set()
for tx in all_bank_txs:
    d = safe_float(tx.get('debit'))
    if d > 0:
        cp = str(tx.get('counterparty', '')).strip().lower().replace(' ', '')
        if cp: bank_payees.add(cp)

unmatched_suppliers = []
for s, info in supplier_stats.items():
    sl = s.lower().replace(' ', '')
    found = any(sl[:4] in bp or bp[:4] in sl for bp in bank_payees)
    if not found: unmatched_suppliers.append((s, info))

unmatched_pur_amt = sum(info['amt'] for _, info in unmatched_suppliers)

# 销项客户 vs 银行收款记录
bank_payers = set()
for tx in all_bank_txs:
    c = safe_float(tx.get('credit'))
    if c > 0:
        cp = str(tx.get('counterparty', '')).strip().lower().replace(' ', '')
        if cp: bank_payers.add(cp)

unmatched_buyers = []
for b, info in buyer_stats.items():
    bl = b.lower().replace(' ', '')
    found = any(bl[:4] in bp or bp[:4] in bl for bp in bank_payers)
    if not found: unmatched_buyers.append((b, info))

# ═══════ 加工费分析 ═══════
proc_invs = []
for i in pur_invs:
    g = str(i.get('goods', '')).lower()
    if any(kw in g for kw in ['加工', '染整', '染色', '印花', '定型', '间棉']):
        proc_invs.append(i)
total_proc = sum(safe_float(i.get('total', i.get('amount'))) for i in proc_invs)

# ═══════ 输出关键数据 ═══════
print("\n" + "=" * 70)
print("关键发现汇总")
print("=" * 70)

print(f"\n【收款来源 TOP10】")
for i, (cp, amt) in enumerate(sorted(credit_by_cp.items(), key=lambda x: -x[1])[:10]):
    pct = amt / max(total_in, 1) * 100
    print(f"  {i+1:2d}. {cp[:35]:35s} {amt:>12,.0f}元 ({pct:.0f}%)")

print(f"\n【付款去向 TOP10】")
for i, (cp, amt) in enumerate(sorted(debit_by_cp.items(), key=lambda x: -x[1])[:10]):
    pct = amt / max(total_out, 1) * 100
    print(f"  {i+1:2d}. {cp[:35]:35s} {amt:>12,.0f}元 ({pct:.0f}%)")

print(f"\n【进项供应商 TOP10】")
for i, (s, info) in enumerate(sorted(supplier_stats.items(), key=lambda x: -x[1]['amt'])[:10]):
    goods = '、'.join(list(info['goods'])[:2])
    print(f"  {i+1:2d}. {s[:25]:25s} {info['cnt']:3d}张 {info['amt']:>12,.0f}元  [{goods[:40]}]")

print(f"\n【销项客户】")
for i, (b, info) in enumerate(sorted(buyer_stats.items(), key=lambda x: -x[1]['amt'])[:10]):
    goods = '、'.join(list(info['goods'])[:2])
    print(f"  {i+1:2d}. {b[:25]:25s} {info['cnt']:3d}张 {info['amt']:>12,.0f}元  [{goods[:40]}]")

# 保存完整数据到JSON
result = {
    "files": 23,
    "company": company_name or "被查单位",
    "period": "2023-06 至 2026-09",
    "pur_invoices": len(pur_invs),
    "sal_invoices": len(sal_invs),
    "bank_txs": len(all_bank_txs),
    "total_pur": round(total_pur, 2),
    "total_sal": round(total_sal, 2),
    "total_bank_in": round(total_in, 2),
    "total_bank_out": round(total_out, 2),
    "gross_margin": round((total_sal - total_pur) / max(total_sal, 1) * 100, 1),
    "suppliers": len(supplier_stats),
    "buyers": len(buyer_stats),
    "unmatched_suppliers": len(unmatched_suppliers),
    "unmatched_pur_amt": round(unmatched_pur_amt, 2),
    "unmatched_buyers": len(unmatched_buyers),
    "proc_invoices": len(proc_invs),
    "total_proc_fee": round(total_proc, 2),
    "credit_top20": {k[:30]: round(v,2) for k,v in sorted(credit_by_cp.items(), key=lambda x:-x[1])[:20]},
    "debit_top20": {k[:30]: round(v,2) for k,v in sorted(debit_by_cp.items(), key=lambda x:-x[1])[:20]},
}

with open('audit_result.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2, default=str)

print(f"\n完整数据已保存到 audit_result.json")
