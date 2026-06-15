"""
7文件类型端到端回归测试
验证 _parse_by_content() 交叉验证引擎对7种真实文件结构的识别准确率
"""
import io
import os
import sys
import json
import traceback
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.utils import get_column_letter

# 导入主模块的解析函数
from main import (
    _parse_by_content,
    _init_trace,
    _get_last_trace,
    _FILE_FINGERPRINTS,
)

def create_xlsx(headers, rows, sheet_name="Sheet1"):
    """创建一个Excel文件并返回BytesIO"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    # 写入数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==================== 7类测试文件定义 ====================

def make_bank_statement():
    """银行流水"""
    headers = ["交易日期", "凭证号", "借方金额", "贷方金额", "余额", "对方户名", "对方账号", "摘要"]
    rows = []
    balance = 100000.00
    for i in range(1, 51):
        debit = round(500 + i * 100, 2) if i % 3 == 0 else 0
        credit = round(300 + i * 80, 2) if i % 3 != 0 else 0
        balance = balance + credit - debit
        rows.append([
            f"2025-0{(i%9)+1:01d}-{(i%28)+1:02d}",
            f"VOUCHER{i:04d}",
            debit,
            credit,
            round(balance, 2),
            f"测试公司{i}",
            f"62220210010{i:06d}",
            f"测试交易{i}"
        ])
    return create_xlsx(headers, rows, "银行流水")

def make_salary():
    """工资表"""
    headers = ["序号", "姓名", "身份证号", "基本工资", "绩效工资", "加班工资", "补贴", "应发合计",
               "养老保险", "医疗保险", "失业保险", "住房公积金", "个税", "实发合计"]
    names = ['张伟', '李娜', '王磊', '刘洋', '陈静', '杨帆', '赵敏', '周杰', '吴芳', '郑强',
             '钱华', '孙丽', '朱峰', '马宁', '胡波', '林涛', '何秀', '郭瑞', '高洁', '罗勇',
             '梁璐', '宋明', '唐颖', '韩冰', '冯俊', '程芳', '曹宇', '袁晨', '邓辉', '许燕']
    rows = []
    for i in range(1, 31):
        name = names[(i-1) % len(names)]
        base = 8000 + i * 200
        perf = 2000 + i * 100
        overtime = i * 50
        allowance = 500
        gross = base + perf + overtime + allowance
        si = round(gross * 0.105, 2)
        tax = round(max(0, gross - 5000 - si) * 0.03, 2)
        net = gross - si - tax
        rows.append([
            i,
            name,
            f"11010119900101{i:04d}",
            base, perf, overtime, allowance, gross,
            round(base * 0.08, 2),
            round(base * 0.02, 2),
            round(base * 0.005, 2),
            round(base * 0.12, 2),
            tax,
            round(net, 2)
        ])
    return create_xlsx(headers, rows, "工资表")

def make_invoice():
    """发票（销项）"""
    headers = ["发票代码", "发票号码", "数电发票号码", "购方名称", "购方税号",
               "销方名称", "销方税号", "开票日期", "货物或应税劳务名称", "金额", "税率", "税额", "价税合计"]
    rows = []
    for i in range(1, 41):
        amount = round(10000 + i * 500, 2)
        tax_rate = 0.13
        tax = round(amount * tax_rate, 2)
        total = round(amount + tax, 2)
        rows.append([
            f"1100{i:04d}",
            f"NO{i:06d}",
            f"",
            f"购买方公司{i}",
            f"91110108MA01ABCD{i:02d}",
            f"销售方公司",
            f"91110108MA01EFGH12",
            f"2025-0{(i%9)+1:01d}-{(i%28)+1:02d}",
            f"商品{i}",
            amount, tax_rate, tax, total
        ])
    return create_xlsx(headers, rows, "发票")

def make_housing_fund():
    """住房公积金"""
    headers = ["姓名", "身份证号", "缴存基数", "单位缴存比例", "个人缴存比例",
               "单位缴存额", "个人缴存额", "合计缴存额"]
    hf_names = ['黄莉', '汪峰', '黎芳', '龙杰', '段芸', '乔宇', '萧敏', '姚刚', '董秀', '潘亮',
                '蔡华', '沈燕', '吕强', '魏婷', '侯建']
    rows = []
    for i in range(1, 21):
        name = hf_names[(i-1) % len(hf_names)]
        base = 8000 + i * 500
        company_ratio = 0.12
        person_ratio = 0.12
        company_pay = round(base * company_ratio, 2)
        person_pay = round(base * person_ratio, 2)
        rows.append([
            name,
            f"11010119900101{i:04d}",
            base,
            company_ratio,
            person_ratio,
            company_pay,
            person_pay,
            round(company_pay + person_pay, 2)
        ])
    return create_xlsx(headers, rows, "公积金")

def make_social_security():
    """社会保险"""
    headers = ["姓名", "身份证号", "缴费基数", "养老保险单位", "养老保险个人",
               "医疗保险单位", "医疗保险个人", "失业保险单位", "失业保险个人",
               "工伤保险单位", "生育保险单位"]
    ss_names = ['刘强', '陈丽', '王军', '张莹', '李华', '赵伟', '孙芳', '周敏', '吴杰', '郑霞',
                '钱勇', '冯兰', '褚明', '卫秀', '蒋峰']
    rows = []
    for i in range(1, 21):
        name = ss_names[(i-1) % len(ss_names)]
        base = 8000 + i * 500
        rows.append([
            name,
            f"11010119900101{i:04d}",
            base,
            round(base * 0.16, 2),
            round(base * 0.08, 2),
            round(base * 0.08, 2),
            round(base * 0.02, 2),
            round(base * 0.007, 2),
            round(base * 0.003, 2),
            round(base * 0.005, 2),
            round(base * 0.008, 2),
        ])
    return create_xlsx(headers, rows, "社保")

def make_trial_balance():
    """科目余额表"""
    headers = ["科目编码", "科目名称", "期初借方", "期初贷方", "本期借方", "本期贷方",
               "期末借方", "期末贷方"]
    rows = []
    accounts = [
        ("1001", "库存现金", 5000, 0, 2000, 1000, 6000, 0),
        ("1002", "银行存款", 500000, 0, 100000, 80000, 520000, 0),
        ("1121", "应收票据", 50000, 0, 0, 0, 50000, 0),
        ("1122", "应收账款", 200000, 0, 50000, 30000, 220000, 0),
        ("1123", "预付账款", 30000, 0, 10000, 5000, 35000, 0),
        ("1221", "其他应收款", 10000, 0, 2000, 0, 12000, 0),
        ("1405", "库存商品", 80000, 0, 20000, 15000, 85000, 0),
        ("1601", "固定资产", 300000, 0, 0, 0, 300000, 0),
        ("1602", "累计折旧", 0, 50000, 0, 5000, 0, 55000),
        ("2001", "短期借款", 0, 100000, 0, 0, 0, 100000),
        ("2201", "应付票据", 0, 20000, 5000, 0, 0, 15000),
        ("2202", "应付账款", 0, 150000, 40000, 60000, 0, 170000),
        ("2203", "预收账款", 0, 30000, 0, 10000, 0, 40000),
        ("2211", "应付职工薪酬", 0, 25000, 20000, 25000, 0, 30000),
        ("2221", "应交税费", 0, 15000, 10000, 12000, 0, 17000),
        ("4001", "实收资本", 0, 1000000, 0, 0, 0, 1000000),
        ("4103", "本年利润", 0, 200000, 0, 0, 0, 200000),
        ("6001", "主营业务收入", 0, 0, 80000, 80000, 0, 0),
        ("6401", "主营业务成本", 0, 0, 50000, 50000, 0, 0),
        ("6601", "销售费用", 0, 0, 30000, 30000, 0, 0),
        ("6602", "管理费用", 0, 0, 25000, 25000, 0, 0),
        ("6603", "财务费用", 0, 0, 2000, 2000, 0, 0),
    ]
    for acc in accounts:
        rows.append(list(acc))
    return create_xlsx(headers, rows, "科目余额表")

def make_contract_list():
    """合同台账"""
    headers = ["合同编号", "合同名称", "合同类型", "甲方", "乙方", "合同金额",
               "签订日期", "到期日期", "状态"]
    rows = []
    for i in range(1, 21):
        amount = 50000 + i * 10000
        rows.append([
            f"HT-2025-{i:04d}",
            f"合同项目{i}",
            "销售合同",
            f"甲方公司{i}",
            "我方公司",
            amount,
            f"2025-0{(i%9)+1:01d}-01",
            f"2026-0{(i%9)+1:01d}-01",
            "执行中"
        ])
    return create_xlsx(headers, rows, "合同台账")

# ==================== 测试执行 ====================

TEST_CASES = [
    ("bank_statement", make_bank_statement, "银行流水"),
    ("salary", make_salary, "工资表"),
    ("invoice", make_invoice, "发票"),
    ("housing_fund", make_housing_fund, "住房公积金"),
    ("social_security", make_social_security, "社会保险"),
    ("trial_balance", make_trial_balance, "科目余额表"),
    ("contract_list", make_contract_list, "合同台账"),
]

def run_test(display_name, make_fn):
    """运行单个文件的解析测试"""
    try:
        buf = make_fn()
        wb = openpyxl.load_workbook(buf, data_only=True)
        names = wb.sheetnames
        def get_sheet(i):
            return wb[names[i]]
        _init_trace(display_name + ".xlsx")  # 初始化诊断追踪
        result = _parse_by_content(names, get_sheet)
        trace = _get_last_trace()
        
        parsed_type = result.get("type") if result else None
        parsed_source = result.get("source") if result else None
        parsed_conf = result.get("confidence") if result else None
        parsed_rows = len(result.get("rows", [])) if result else 0
        
        return {
            "display": display_name,
            "success": True,
            "parsed_type": parsed_type,
            "parsed_source": parsed_source,
            "parsed_confidence": parsed_conf,
            "parsed_rows": parsed_rows,
            "trace": {
                "final_decision": trace.get("final_decision", {}),
                "cross_validation": trace.get("cross_validation", {}),
                "structure_phase": trace.get("structure_phase", {}),
                "keyword_phase": trace.get("keyword_phase", {}),
                "suggestions": trace.get("suggestions", []),
            }
        }
    except Exception as e:
        return {
            "display": display_name,
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc(),
        }

def run_all_tests():
    results = []
    passed = 0
    failed = 0
    
    print("=" * 70)
    print("7文件类型端到端回归测试 — 交叉验证引擎")
    print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    for expected_type, make_fn, display_name in TEST_CASES:
        print(f"\n--- {display_name} (期望: {expected_type}) ---")
        r = run_test(display_name, make_fn)
        results.append((expected_type, r))
        
        if r["success"]:
            pt = r["parsed_type"]
            ps = r["parsed_source"]
            pc = r["parsed_confidence"]
            pr = r["parsed_rows"]
            
            # 类型映射
            type_map = {
                "bank_statement": "bank_transaction|bank_statement",
                "invoice": "invoice|sales_invoice|purchase_invoice|invoice_universal",
                "social_security": "social_security",
                "housing_fund": "housing_fund",
                "salary": "salary",
                "trial_balance": "trial_balance",
                "contract_list": "contract|contract_list",
            }
            accepted = type_map.get(expected_type, expected_type).split("|")
            
            is_correct = pt in accepted
            status = "✅" if is_correct else "❌"
            
            if is_correct:
                passed += 1
            else:
                failed += 1
            
            print(f"  识别: {pt} (来源:{ps}, 置信度:{pc if pc is not None else 'N/A'} 行数:{pr}) {status}")
            
            # 显示裁决详情
            cv = r["trace"]["cross_validation"]
            if cv:
                print(f"  裁决: {cv.get('winner', '?')} — {cv.get('reason', '')[:80]}")
            
            # 检查诊断建议
            suggestions = r["trace"]["suggestions"]
            if suggestions:
                print(f"  ⚠ 诊断建议 ({len(suggestions)}条):")
                for s in suggestions[:3]:
                    print(f"    - {s.get('issue', '')[:60]}")
        else:
            print(f"  ❌ 异常: {r['error']}")
            failed += 1
    
    # 总结
    print("\n" + "=" * 70)
    print(f"测试结果: {passed}/{len(TEST_CASES)} 通过, {failed} 失败")
    print("=" * 70)
    
    return passed == len(TEST_CASES)

if __name__ == "__main__":
    ok = run_all_tests()
    sys.exit(0 if ok else 1)
